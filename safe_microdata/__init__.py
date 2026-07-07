"""
ECB SAFE Microdata — dlt sources
==================================
Two sources:

safe_microdata_source
---------------------
Downloads the ECB SAFE microdata zip, extracts every CSV, and yields rows
for the dlt pipeline to load into MotherDuck (dataset: raw).

Incremental behaviour: ETag/Last-Modified stored in dlt state; skips load
if headers unchanged from the last successful run.

Config (via .dlt/config.toml  [sources.safe_microdata_source])
--------------------------------------------------------------
  zip_url              Primary download URL for the zip file.
  fallback_scrape_url  Page to scrape for a .zip link if zip_url returns 404.
  request_timeout      HTTP timeout in seconds (default 60).

safe_annex_source
-----------------
Downloads the ECB SAFE questionnaire annex XLSX once and yields two resources:
  - ref_safe__annex       wide, one row per spreadsheet row (legacy shape).
  - ref_safe__annex_long  unpivoted, one row per (question_item, element,
                          wave_label), non-blank cells only. dbt models build
                          on this long table instead of every consumer running
                          its own information_schema.columns + COALESCE query.

Both land in dataset main_safe. Always a full refresh (write_disposition=
"replace") since the annex is small (~1 300 rows) and changes infrequently.

Exits 1 if the URL is unreachable or the file cannot be parsed — CI will
catch this non-zero exit and open a GitHub Issue.
"""

from __future__ import annotations
import io
import re
import sys
import csv
import logging
import tempfile
import zipfile
from pathlib import Path
from typing import Iterator
from urllib.parse import urljoin

import dlt
import openpyxl
import requests
from bs4 import BeautifulSoup

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# HTTP helpers
# ---------------------------------------------------------------------------

def _make_session() -> requests.Session:
    session = requests.Session()
    session.headers["User-Agent"] = (
        "safe-ai-dlt/1.0 (research pipeline; contact: lubos.pernis@gmail.com)"
    )
    return session


def _scrape_zip_url(session: requests.Session, scrape_url: str, timeout: int) -> str:
    """Scrape *scrape_url* and return the first absolute .zip href found."""
    logger.info("Scraping fallback page: %s", scrape_url)
    resp = session.get(scrape_url, timeout=timeout)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")
    for tag in soup.find_all("a", href=True):
        href: str = tag["href"]
        if href.lower().endswith(".zip"):
            url = urljoin(scrape_url, href)
            logger.info("Found zip link: %s", url)
            return url
    raise RuntimeError(f"No .zip link found on fallback page: {scrape_url}")


# ---------------------------------------------------------------------------
# dlt source / resource
# ---------------------------------------------------------------------------

@dlt.source
def safe_microdata_source(
    zip_url: str = dlt.config.value,
    fallback_scrape_url: str = dlt.config.value,
    request_timeout: int = dlt.config.value,
) -> dlt.sources.DltSource:
    """dlt source for the ECB SAFE microdata zip."""
    return _safe_microdata(zip_url, fallback_scrape_url, request_timeout)


@dlt.resource(
    name="safe_microdata",
    write_disposition="replace",
)
def _safe_microdata(
    zip_url: str,
    fallback_scrape_url: str,
    request_timeout: int,
) -> Iterator[dict]:
    session = _make_session()
    state = dlt.current.source_state()

    # ------------------------------------------------------------------ #
    # 1. HEAD request — resolve URL and fetch cache headers               #
    # ------------------------------------------------------------------ #
    logger.info("HEAD %s", zip_url)
    resp = session.head(zip_url, timeout=request_timeout, allow_redirects=True)

    if resp.status_code == 404:
        logger.warning("Primary URL returned 404 — scraping fallback page.")
        zip_url = _scrape_zip_url(session, fallback_scrape_url, request_timeout)
        resp = session.head(zip_url, timeout=request_timeout, allow_redirects=True)
        resp.raise_for_status()
    else:
        resp.raise_for_status()

    etag: str | None = resp.headers.get("ETag")
    last_modified: str | None = resp.headers.get("Last-Modified")
    logger.info("ETag=%s  Last-Modified=%s", etag, last_modified)

    # ------------------------------------------------------------------ #
    # 2. Watermark check — skip if nothing has changed                    #
    # ------------------------------------------------------------------ #
    if etag and etag == state.get("etag"):
        logger.info("ETag unchanged — no new data, skipping load.")
        sys.exit(0)
        return
    if not etag and last_modified and last_modified == state.get("last_modified"):
        logger.info("Last-Modified unchanged — no new data, skipping load.")
        sys.exit(0)
        return

    # ------------------------------------------------------------------ #
    # 3. Download zip and extract CSVs                                    #
    # ------------------------------------------------------------------ #
    rows: list[dict] = []

    with tempfile.TemporaryDirectory() as tmp_dir:
        zip_path = Path(tmp_dir) / "safe_microdata.zip"

        logger.info("Downloading zip from %s …", zip_url)
        with session.get(zip_url, timeout=request_timeout, stream=True) as dl:
            dl.raise_for_status()
            with zip_path.open("wb") as fh:
                for chunk in dl.iter_content(chunk_size=1 << 20):
                    fh.write(chunk)
        size_mb = zip_path.stat().st_size / (1 << 20)
        logger.info("Download complete — %.1f MiB", size_mb)

        with zipfile.ZipFile(zip_path, "r") as zf:
            csv_names = [n for n in zf.namelist() if n.lower().endswith(".csv")]
            if not csv_names:
                raise RuntimeError("No CSV files found inside the downloaded zip.")
            logger.info("Extracting %d CSV file(s): %s", len(csv_names), csv_names)
            zf.extractall(tmp_dir)

        for csv_name in csv_names:
            csv_path = Path(tmp_dir) / csv_name
            logger.info("Reading %s …", csv_name)
            with csv_path.open("r", encoding="utf-8", errors="replace") as fh:
                for row in csv.DictReader(fh):
                    rows.append(dict(row))

    logger.info("Yielding %d rows to dlt.", len(rows))

    # ------------------------------------------------------------------ #
    # 4. Yield rows — outside the temp dir context so it is already gone  #
    # ------------------------------------------------------------------ #
    yield from rows

    # ------------------------------------------------------------------ #
    # 5. Persist watermark after successful yield                         #
    # ------------------------------------------------------------------ #
    if etag:
        state["etag"] = etag
    if last_modified:
        state["last_modified"] = last_modified
    logger.info("State updated — etag=%s  last_modified=%s", etag, last_modified)


# ---------------------------------------------------------------------------
# ECB SAFE annex source
# ---------------------------------------------------------------------------

ANNEX_URL = "https://www.ecb.europa.eu/stats/pdf/surveys/sme/Annex_3.en.xlsx"
_ANNEX_MIN_ROWS = 100


def _safe_col_name(raw: str | None, index: int) -> str:
    """Sanitise a header cell into a valid SQL/dlt column name."""
    if not raw or not str(raw).strip():
        return f"col_{index}"
    name = str(raw).strip()
    name = re.sub(r"[^a-zA-Z0-9]", "_", name)
    name = re.sub(r"_+", "_", name).strip("_").lower()
    return name or f"col_{index}"


def _download_and_parse_annex() -> tuple[list[str], list[list[str]]]:
    """Download the annex XLSX and return (column_names, padded_data_rows).

    Shared by both the wide (_safe_annex) and long (_safe_annex_long) resources
    so the XLSX is only downloaded/parsed once per pipeline run.
    """
    session = _make_session()

    logger.info("Downloading annex from %s …", ANNEX_URL)
    try:
        resp = session.get(ANNEX_URL, timeout=60)
        resp.raise_for_status()
    except Exception as exc:
        logger.error("Could not download annex: %s", exc)
        sys.exit(1)

    logger.info("Parsing XLSX (%.1f KiB) …", len(resp.content) / 1024)
    try:
        wb = openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)
        sheet_name = "Questionnaire" if "Questionnaire" in wb.sheetnames else wb.sheetnames[0]
        sheet = wb[sheet_name]
        all_rows = list(sheet.values)
    except Exception as exc:
        logger.error("Could not parse annex XLSX: %s", exc)
        sys.exit(1)

    if len(all_rows) < _ANNEX_MIN_ROWS:
        logger.error(
            "Annex has only %d rows (expected %d+) — sheet structure may have changed.",
            len(all_rows), _ANNEX_MIN_ROWS,
        )
        sys.exit(1)

    # Row 0 = round-number header; Row 1 = column labels; Row 2+ = data
    header_row = all_rows[1]
    data_rows = all_rows[2:]

    col_names = [_safe_col_name(c, i) for i, c in enumerate(header_row)]
    n_cols = len(col_names)

    # De-duplicate column names (some wave columns share sanitised names)
    seen: dict[str, int] = {}
    deduped: list[str] = []
    for name in col_names:
        if name in seen:
            seen[name] += 1
            deduped.append(f"{name}_{seen[name]}")
        else:
            seen[name] = 0
            deduped.append(name)
    col_names = deduped

    padded_rows: list[list[str]] = []
    for row in data_rows:
        padded = [str(c) if c is not None else "" for c in row]
        padded = (padded + [""] * n_cols)[:n_cols]
        if any(v.strip() for v in padded):
            padded_rows.append(padded)

    logger.info(
        "Annex: parsed %d rows (%d columns) from sheet '%s'", len(padded_rows), n_cols, sheet_name,
    )
    return col_names, padded_rows


@dlt.source
def safe_annex_source() -> dlt.sources.DltSource:
    """dlt source: ECB SAFE questionnaire annex XLSX → main_safe.ref_safe__annex
    (wide, one row per sheet row) and main_safe.ref_safe__annex_long (unpivoted,
    one row per question_item × element × wave, non-blank cells only)."""
    col_names, padded_rows = _download_and_parse_annex()
    return (
        _safe_annex(col_names, padded_rows),
        _safe_annex_long(col_names, padded_rows),
    )


@dlt.resource(
    name="ref_safe__annex",
    write_disposition="replace",
)
def _safe_annex(col_names: list[str], padded_rows: list[list[str]]) -> Iterator[dict]:
    """Yield one dict per annex sheet row, in its original wide shape."""
    n_yielded = 0
    for padded in padded_rows:
        yield dict(zip(col_names, padded))
        n_yielded += 1
    logger.info("Annex (wide): yielded %d rows.", n_yielded)


# Fixed, non-wave metadata columns carried on every long-format row for filtering/joins.
_ANNEX_LONG_META_COLS = ("element", "question_item", "answer", "sample", "notes")


@dlt.resource(
    name="ref_safe__annex_long",
    write_disposition="replace",
)
def _safe_annex_long(col_names: list[str], padded_rows: list[list[str]]) -> Iterator[dict]:
    """Unpivot the wave-labelled columns (safe_2024q1, safe_2024q2, ...) into a long
    table: one row per (question_item, element, wave_label), non-blank cells only.

    This removes the need for downstream consumers to run
    `information_schema.columns` introspection + COALESCE(NULLIF(...)) over the wide
    annex table — dbt models can build directly on this long table with plain SQL.
    """
    col_index = {name: i for i, name in enumerate(col_names)}
    wave_cols = [c for c in col_names if c.startswith("safe_")]
    meta_cols = [c for c in _ANNEX_LONG_META_COLS if c in col_index]

    n_yielded = 0
    for padded in padded_rows:
        meta = {c: padded[col_index[c]] for c in meta_cols}
        for wave_col in wave_cols:
            text = padded[col_index[wave_col]]
            if not text or not text.strip():
                continue
            yield {**meta, "wave_label": wave_col, "text": text}
            n_yielded += 1

    logger.info("Annex (long): yielded %d rows across %d wave columns.", n_yielded, len(wave_cols))
